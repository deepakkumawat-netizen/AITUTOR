"""Parse GRADE Wise TOC.xlsx into a structured RAG knowledge base.

Sheet layouts vary:
- Grade 1-2: Subject in FIRST row only (then forward-filled), Unit rows mixed in
- Grade 5: Subject repeated every row, no Stream column
- Grade 6+: Has extra 'Stream / Module' column
"""
import json
from openpyxl import load_workbook

SRC = r'C:\Users\DEEPAK\Downloads\GRADE Wise TOC.xlsx'
OUT = r'C:\Users\DEEPAK\Desktop\ai tutor\backend\cbse_toc.json'

SUBJECT_KEYWORDS = {
    "MATHS", "MATHEMATICS", "SCIENCE", "ENGLISH", "HINDI", "EVS",
    "SOCIAL", "SOCIAL SCIENCE", "SOCIAL STUDIES", "PHYSICS", "CHEMISTRY",
    "BIOLOGY", "COMPUTER", "COMPUTER SCIENCE", "GEOGRAPHY", "HISTORY",
    "CIVICS", "ECONOMICS", "ACCOUNTANCY", "BUSINESS STUDIES",
    "POLITICAL SCIENCE", "SANSKRIT", "INFORMATICS PRACTICES",
}


def normalize_subject(s: str) -> str:
    s = s.strip()
    up = s.upper()
    if up in ("MATHS", "MATHEMATICS"):
        return "Maths"
    if up == "SOCIAL":
        return "Social Science"
    if up == "EVS":
        return "EVS"
    return s.title()


def is_subject_only_row(cells):
    """True if row is just a subject label (other cells empty)."""
    non_empty = [c for c in cells if c]
    if len(non_empty) != 1:
        return False
    return non_empty[0].strip().upper() in SUBJECT_KEYWORDS


def is_unit_row(cells):
    joined = " ".join(cells).lower()
    if "unit" not in joined:
        return False
    # Unit rows have only 1-2 non-empty cells (the unit name)
    non_empty = [c for c in cells if c]
    return len(non_empty) <= 2


def detect_columns(header_cells):
    """Map column index to its role. Returns dict with keys:
    subject, stream, chapter_no, title, concepts."""
    cols = {}
    for i, c in enumerate(header_cells):
        low = c.lower()
        if "subject" in low:
            cols["subject"] = i
        elif "stream" in low or "module" in low:
            cols["stream"] = i
        elif "no." in low or "no" == low.strip() or "number" in low or "chapter / unit" in low or "chapter/unit" in low:
            cols["chapter_no"] = i
        elif "topic name" in low or "chapter name" in low or "topic" in low or "chapter / topic" in low:
            cols["title"] = i
        elif "concept" in low or "description" in low:
            cols["concepts"] = i
    return cols


def find_header_row(rows):
    """Find the row that contains column headers like 'Chapter' and 'Topic Name'."""
    for i, row in enumerate(rows[:8]):
        cells = [str(c or "").strip() for c in row]
        joined = " ".join(cells).lower()
        if ("chapter" in joined and ("topic" in joined or "concept" in joined)) and any("no" in c.lower() for c in cells):
            return i
    return None


def parse_sheet(ws, grade_label):
    rows = list(ws.iter_rows(values_only=True))
    rows_str = [[(str(c).strip() if c is not None else "") for c in row] for row in rows]

    # Some sheets (Grade 1-2) have subject AS the header (e.g. "MATHS | Chapter | Topic ...")
    # Treat the first such row as both subject AND header.
    header_idx = find_header_row(rows_str)
    if header_idx is None:
        # Try: row 0 has subject in col 0 and headers in cols 1+
        if rows_str and rows_str[0][0].upper() in SUBJECT_KEYWORDS:
            header_idx = 0
        else:
            return {}

    header = rows_str[header_idx]
    cols = detect_columns(header)

    # If header row has a subject keyword in col 0 (Grade 1 style), use it
    inline_subject = None
    if header[0] and header[0].strip().upper() in SUBJECT_KEYWORDS:
        inline_subject = normalize_subject(header[0])

    result = {}
    current_subject = inline_subject
    current_unit = None

    for row in rows_str[header_idx + 1:]:
        # All blank?
        if not any(row):
            current_unit = None
            continue

        # Header row repeating inside the sheet — skip
        if header_idx is not None:
            low = " ".join(row).lower()
            if "chapter / topic" in low or "chapter name" in low or ("chapter" in low and "concept" in low and "no" in low):
                continue

        # Subject-only row (e.g. "ENGLISH" alone in a cell)
        if is_subject_only_row(row):
            non_empty = [c for c in row if c][0]
            current_subject = normalize_subject(non_empty)
            current_unit = None
            continue

        # Unit row
        if is_unit_row(row):
            non_empty = [c for c in row if c]
            current_unit = " ".join(non_empty).strip()
            continue

        # Get values from the column-mapped positions
        def at(key):
            idx = cols.get(key)
            if idx is None or idx >= len(row):
                return ""
            return row[idx]

        subj_in_row = at("subject")
        if subj_in_row and subj_in_row.upper() in SUBJECT_KEYWORDS:
            current_subject = normalize_subject(subj_in_row)

        ch_no = at("chapter_no")
        title = at("title")
        concepts = at("concepts")
        stream = at("stream") if "stream" in cols else ""

        if not (ch_no and title):
            continue
        if not current_subject:
            continue

        entry = {"ch": ch_no, "title": title}
        if current_unit:
            entry["unit"] = current_unit
        if stream:
            entry["stream"] = stream
        if concepts:
            entry["concepts"] = concepts

        result.setdefault(current_subject, []).append(entry)

    return result


wb = load_workbook(SRC, data_only=True)
curriculum = {}

for sheet_name in wb.sheetnames:
    grade_num = sheet_name.split()[-1]
    grade_label = f"Grade {grade_num}"
    sheet_data = parse_sheet(wb[sheet_name], grade_label)
    if sheet_data:
        curriculum[grade_label] = sheet_data

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(curriculum, f, ensure_ascii=False, indent=2)

print(f"Wrote {OUT}")
for g in curriculum:
    subjects = list(curriculum[g].keys())
    chap_count = sum(len(curriculum[g][s]) for s in subjects)
    print(f"  {g}: {len(subjects)} subjects ({', '.join(subjects)}), {chap_count} chapters")
